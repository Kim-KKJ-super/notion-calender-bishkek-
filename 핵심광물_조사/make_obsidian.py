# -*- coding: utf-8 -*-
"""
옵시디언 보관함에 넣을 폴더를 생성한다.

- 중간보고 2차, 노츠메일 문안: 그대로 노트로 이관(프론트매터·위키링크 추가)
- 별첨 2 엑셀: 옵시디언에서 내용이 보이지 않으므로, 시트별 마크다운 노트로 변환하고
  원본 xlsx는 `첨부/` 폴더에 함께 넣는다.
"""
import os, re, shutil
from openpyxl import load_workbook

ROOT = "/home/user/notion-calender-bishkek-"
OUT = os.path.join(ROOT, "obsidian_export", "핵심광물 조사 (2026.9)")
ATT = os.path.join(OUT, "첨부")
XLSX = os.path.join(ROOT, "핵심광물_조사/산출물/[주키르기즈대사관] 핵심광물 관련 현황(별첨2)_20260905_v1.xlsx")

shutil.rmtree(os.path.join(ROOT, "obsidian_export"), ignore_errors=True)
os.makedirs(ATT, exist_ok=True)

TAGS = "  - 업무/핵심광물\n  - 주재국/키르기스스탄\n  - 본부보고"

def fm(title, extra=""):
    return (f"---\ntitle: {title}\ndate: 2026-09-05\n제출기한: 2026-09-15\n"
            f"업무: 핵심광물 관련 현황 조사\n발신: 장관(경제안보외교과장)\n"
            f"상태: 1차본(교차검증 전)\ntags:\n{TAGS}\n{extra}---\n\n")

# ---------------------------------------------------------------- 00 개요(MOC)
index = fm("핵심광물 조사 개요", "cssclasses:\n  - wide\n")
index += """# 핵심광물 관련 현황 조사 (2026.9)

> [!info] 문서 성격
> 본부 「핵심광물 관련 현황 조사」(2026.9.3. 장관/경제안보외교과) 대응 자료.
> **제출기한 2026.9.15.(화)**

> [!warning] 1차본이며 교차검증 전입니다
> 수치는 2026.9.5. 기준 공개자료(USGS, 영국 정부 Growth Gateway, EITI, UN Comtrade,
> 주재국 Geoportal, 현지 언론) 및 공관 보유자료에 기반합니다.
> 원문 대조가 필요한 항목이 남아 있으며, 교차검증 9.12.~13. / 최종 점검 9.14. 예정입니다.

## 노트

- [[01 중간보고 2차]] — 진행률, 확인된 사실, 자료 충돌 처리, 미확인 항목, 결정 요청
- [[02 노츠메일 문안]] — 담당자 앞 송부 문안 초안
- [[03 별첨2 - 1. 핵심광물 광종별 현황]] — 38종 + 참고 5종
- [[04 별첨2 - 2.1 우리 정부와의 협력 현황]]
- [[05 별첨2 - 2.2 우리 기업 진출 현황]]
- [[06 별첨2 - 3.1 제3국과의 협력 현황]]
- [[07 별첨2 - 3.2 주재국 내 외국기업 진출 현황]]

## 제출 파일

실제 본부 제출본은 아래 엑셀입니다. 03~07 노트는 옵시디언에서 읽기 위한 변환본이며,
**내용을 고칠 때는 엑셀 원본을 고쳐야 합니다.**

![[[주키르기즈대사관] 핵심광물 관련 현황(별첨2)_20260905_v1.xlsx]]

## 핵심 결론

- 우리 정부 지정 **38종 중 부존·채굴·제련이 모두 확인되는 광종은 안티모니 1종**뿐입니다.
- **구리**는 채굴·정광 수출까지만 확인됩니다(주재국 내 제련시설 부재).
- **희토류**는 부존만 확인되며 **1991년 이후 생산이 없습니다**. JORC 기준 자원량 평가·
  타당성조사·확정 투자사업이 모두 부재합니다.
- 38종 중 **10종**은 공개자료상 부존 여부 자체가 확인되지 않습니다.
- **한-키 협력은 전부 MOU·협의 단계**이며, 우리 기업의 광산 지분투자·장기공급계약·
  제련사업은 확인되지 않습니다.
- **국유화 리스크**가 최대 특기사항입니다 — 쿰토르 완전 국유화(2022), 희토류·베릴륨
  라이선스의 국영기업 이관(2024.1.), 채굴허가 199건 취소(2025년 상반기).

## 표현상 유의사항

> [!caution]
> - MOU·협의를 투자·생산 실적으로 쓰지 말 것
> - 포스코는 언론 보도상 "협의 중"이며 "진출"이 아님
> - 안티모니 對중국 수출액은 **중국 해관총서 기준**(주재국 관세청 통계 아님)
> - 매장량은 대부분 소련식(GKZ) 분류 추정 — JORC와 직접 비교 불가
> - 중앙아 지역 합계 매장 비중(망간 39% 등)을 주재국 수치로 전용 금지
> - 중국의 직접투자·정책금융 차관·개발권 대가 인프라 투자를 합산 표현 금지

## 관련 Notion

열람용 사본이 Notion에도 있습니다 — `🔒 대사 개인 공간 / 대사전용 참고자료 /
핵심광물 관련 현황 조사(2026.9) — 본부 제출자료`
(시트별 데이터베이스 5개, 총 64행). 수정 시 엑셀·Notion·본 노트를 함께 갱신해야 합니다.
"""
open(os.path.join(OUT, "00 핵심광물 조사 개요.md"), "w", encoding="utf-8").write(index)

# ---------------------------------------------------- 01/02 기존 마크다운 이관
def port(src, dst, title, back=True):
    body = open(os.path.join(ROOT, "핵심광물_조사", src), encoding="utf-8").read()
    # 최상단 H1은 프론트매터 title과 중복되므로 유지하되 앞에 백링크를 넣는다
    head = fm(title)
    if back:
        head += "> 상위 노트: [[00 핵심광물 조사 개요]]\n\n"
    open(os.path.join(OUT, dst), "w", encoding="utf-8").write(head + body)

port("중간보고_2차_20260905.md", "01 중간보고 2차.md", "중간보고 2차 (2026.9.5.)")
port("노츠메일_문안.md", "02 노츠메일 문안.md", "노츠메일 송부 문안 초안")

# --------------------------------------------------------- 03~07 엑셀 → 노트
wb = load_workbook(XLSX)
SHEETS = [
    ("1. 핵심광물 광종별 현황", "03 별첨2 - 1. 핵심광물 광종별 현황.md", 3, 2),
    ("2.1.우리 정부와의 협력 현황", "04 별첨2 - 2.1 우리 정부와의 협력 현황.md", 2, 4),
    ("2.2. 우리 기업 진출 현황", "05 별첨2 - 2.2 우리 기업 진출 현황.md", 2, 6),
    ("3.1. 제3국과의 협력 현황", "06 별첨2 - 3.1 제3국과의 협력 현황.md", 2, 5),
    ("3.2. 주재국 내 외국기업 진출 현황", "07 별첨2 - 3.2 주재국 내 외국기업 진출 현황.md", 2, 7),
]

def clean(v):
    if v is None:
        return ""
    return re.sub(r"\s*\n\s*", " ", str(v)).strip()

for sheet, fname, start, title_col in SHEETS:
    ws = wb[sheet]
    hdr = [clean(c.value) for c in ws[1]]
    out = fm(f"별첨2 — {sheet}")
    out += f"> 상위 노트: [[00 핵심광물 조사 개요]]\n\n# 별첨 2 — {sheet}\n\n"
    out += ("> [!note] 이 노트는 제출용 엑셀의 변환본입니다\n"
            "> 열 이름은 본부 양식 그대로이며, 수정은 엑셀 원본에서 하십시오.\n"
            f"> 원본 시트명: `{sheet}`\n\n")
    if start == 3:
        out += ("> [!abstract] 기입안내(양식 2행, 원본 보존)\n"
                "> A열 `국가명` / B열 `광종명` / C열 `해당시 작성`\n\n")
    for row in ws.iter_rows(min_row=start, max_row=ws.max_row):
        vals = [clean(c.value) for c in row]
        if not any(vals):
            continue
        out += f"## {vals[title_col - 1]}\n\n"
        for h, v in zip(hdr, vals):
            if not h or not v:
                continue
            out += f"- **{h}**\n\t- {v}\n"
        out += "\n"
    open(os.path.join(OUT, fname), "w", encoding="utf-8").write(out)

shutil.copy2(XLSX, ATT)
print("생성 완료")
for r, d, f in os.walk(os.path.join(ROOT, "obsidian_export")):
    for x in sorted(f):
        p = os.path.join(r, x)
        print(f"  {os.path.relpath(p, os.path.join(ROOT, 'obsidian_export'))}  ({os.path.getsize(p):,} bytes)")
